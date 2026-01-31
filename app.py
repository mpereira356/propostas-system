"""
Aplicação Flask para gerenciamento de propostas comerciais
"""
import os
import io
import re
import threading
from queue import Queue
from datetime import datetime, timedelta, date
from math import ceil
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, send_from_directory
from werkzeug.utils import secure_filename
from models import db, Proposta, ItemProposta, Cliente, Setor, Regiao, Visita, Contato, Equipamento, init_db
from pdf_reader import PropostaExtractor
from sqlalchemy import text, func, literal, case
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configurações
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}

app = Flask(__name__)
app.config['SECRET_KEY'] = 'sua-chave-secreta-aqui-mude-em-producao'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
app.config['MAX_FILES_PER_UPLOAD'] = 10
app.config['BACKFILL_MAX_PER_REQUEST'] = 0

upload_queue = Queue()
upload_lock = threading.Lock()
upload_total = 0
upload_done = 0
reprocess_status = {'running': False, 'total': 0, 'done': 0, 'errors': 0}


def process_pdf(filepath, filename_original, filename):
    try:
        extractor = PropostaExtractor(filepath)
        dados = extractor.extract_all()
        base_id = extract_base_id_from_filename(filename_original or filename)
        if not dados or (not dados.get('id_proposta') and not base_id):
            print(f"Falha ao extrair dados do PDF: {filename_original}")
            return

        cod_vendedor = extract_cod_from_filename(filename_original)
        data_emissao_date = parse_date_br(dados.get('data_emissao'))
        data_vencimento = compute_data_vencimento(dados.get('validade'), data_emissao_date)
        if not data_vencimento and data_emissao_date:
            data_vencimento = data_emissao_date + timedelta(days=30)

        proposta_existente = Proposta.query.filter_by(nome_arquivo_pdf=filename).first()

        if proposta_existente:
            print(f"Proposta já importada: {dados['id_proposta']}")
            return

        versao = extract_version_from_filename(filename_original or filename)
        id_proposta = ensure_unique_id_proposta(dados.get('id_proposta'), base_id, filename_original or filename)

        razao_social = dados.get('razao_social')
        razao_social = razao_social.upper() if isinstance(razao_social, str) and razao_social.strip() else razao_social
        tipo = dados.get('tipo') or extract_tipo_from_filename(filename_original or filename)

        proposta = Proposta(
            razao_social=razao_social,
            nome_fantasia=dados.get('nome_fantasia'),
            id_proposta=id_proposta or dados.get('id_proposta'),
            data_emissao=dados.get('data_emissao'),
            validade=dados.get('validade'),
            cnpj=dados.get('cnpj'),
            telefone=dados.get('telefone'),
            celular=dados.get('celular'),
            email=dados.get('email'),
            pessoa_contato=dados.get('pessoa_contato'),
            descricao_item=dados.get('descricao_item'),
            quantidade=dados.get('quantidade'),
            valor_total=dados.get('valor_total'),
            nome_arquivo_pdf=filename,
            cod_vendedor=cod_vendedor,
            data_vencimento=data_vencimento,
            instalacao_status=dados.get('instalacao_status'),
            qualificacoes_status=dados.get('qualificacoes_status'),
            treinamento_status=dados.get('treinamento_status'),
            garantia_resumo=dados.get('garantia_resumo'),
            garantia_texto=dados.get('garantia_texto'),
            tipo=tipo,
            observacoes='Em negociação'
        )
        proposta.id_proposta_base = base_id
        proposta.versao = versao

        db.session.add(proposta)
        db.session.flush()

        for item_data in dados.get('itens', []):
            item = ItemProposta(
                proposta_id=proposta.id,
                numero=item_data.get('numero'),
                descricao=item_data.get('descricao'),
                quantidade=item_data.get('quantidade'),
                valor_unitario=item_data.get('valor_unitario'),
                valor_total=item_data.get('valor_total')
            )
            db.session.add(item)

        cnpj = dados.get('cnpj')
        if cnpj:
            cnpj_normalizado = normalize_cnpj(cnpj)
            cliente_existente = Cliente.query.filter(
                (Cliente.cnpj == cnpj) | (Cliente.cnpj_normalizado == cnpj_normalizado)
            ).first()
            if not cliente_existente:
                nome_cliente = dados.get('razao_social') or dados.get('nome_fantasia') or 'Cliente sem nome'
                cliente = Cliente(
                    nome=nome_cliente,
                    cnpj=cnpj,
                    cnpj_normalizado=cnpj_normalizado
                )
                db.session.add(cliente)

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao processar PDF ({filename_original}): {e}")


def upload_worker():
    while True:
        item = upload_queue.get()
        if item is None:
            break
        filepath, filename_original, filename = item
        with app.app_context():
            process_pdf(filepath, filename_original, filename)
        with upload_lock:
            global upload_done
            upload_done += 1
        upload_queue.task_done()


worker_thread = threading.Thread(target=upload_worker, daemon=True)
worker_thread.start()

def parse_date_br(date_str):
    """Converte data no formato dd/mm/aaaa para datetime.date."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").date()
    except Exception:
        return None


def parse_date_iso(date_str):
    """Converte data no formato aaaa-mm-dd para datetime.date."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return None


def normalize_cnpj(cnpj):
    """Normaliza CNPJ removendo caracteres não numéricos."""
    if not cnpj:
        return None
    return re.sub(r'\D', '', cnpj)


def extract_cod_from_filename(filename):
    """Extrai o código do vendedor do nome do arquivo."""
    if not filename:
        return None
    match = re.search(r'cod\s*([A-Za-z0-9]+)', filename, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def extract_base_id_from_filename(filename):
    """Extrai os 3 primeiros dígitos do nome do arquivo como base da proposta."""
    if not filename:
        return None
    name = os.path.basename(filename)
    match = re.match(r'^\s*(\d{3})', name)
    if not match:
        return None
    return match.group(1)


def extract_version_from_filename(filename):
    """Extrai um sufixo curto de versão após os 3 dígitos iniciais."""
    if not filename:
        return ''
    name = os.path.splitext(os.path.basename(filename))[0]
    match = re.match(r'^\s*\d{3}(.*)$', name)
    if not match:
        return ''
    tail = match.group(1)
    parts = re.split(r'[\s_.-]+', tail)
    for part in parts:
        cleaned = re.sub(r'[^A-Za-z0-9]', '', part)
        if cleaned:
            return cleaned[:5].upper()
    return ''


def ensure_unique_id_proposta(candidate_id, base_id, filename):
    """Garante que o id_proposta seja único; usa base/sufixo do arquivo se necessário."""
    candidate = candidate_id or base_id
    if not candidate:
        return None
    if not Proposta.query.filter_by(id_proposta=candidate).first():
        return candidate

    suffix = extract_version_from_filename(filename)
    if base_id and suffix:
        candidate = f"{base_id}-{suffix}"
        if not Proposta.query.filter_by(id_proposta=candidate).first():
            return candidate

    base = base_id or re.sub(r'\s+', '', (candidate_id or 'PROP'))[:10]
    base = re.sub(r'[^A-Za-z0-9-]', '', base) or 'PROP'
    name = os.path.splitext(os.path.basename(filename or ''))[0]
    name = re.sub(r'[^A-Za-z0-9-]', '', name)[:10]
    seed = f"{base}-{name}" if name else base
    seed = seed[:45]
    attempt = seed
    counter = 1
    while Proposta.query.filter_by(id_proposta=attempt).first():
        attempt = f"{seed}-{counter}"
        counter += 1
    return attempt


def extract_tipo_from_filename(filename):
    """Detecta tipo pelo nome do arquivo quando o PDF não traz a marca."""
    if not filename:
        return None
    name = os.path.basename(filename)
    if re.search(r'(^|[^A-Za-z0-9])MP\s*BIOS([^A-Za-z0-9]|$)', name, re.IGNORECASE) or re.search(r'(^|[^A-Za-z0-9])MPBIOS([^A-Za-z0-9]|$)', name, re.IGNORECASE):
        return 'Serviço'
    if re.search(r'(^|[^A-Za-z0-9])BAUMER([^A-Za-z0-9]|$)', name, re.IGNORECASE):
        return 'Produto'
    return None


def compute_data_vencimento(validade, data_emissao_date):
    """Calcula data de vencimento a partir da validade."""
    if validade:
        match = re.search(r'(\d{2})\s*[./]\s*(\d{2})\s*[./]\s*(\d{2,4})', validade)
        if match:
            dia, mes, ano = match.groups()
            if len(ano) == 2:
                ano = '20' + ano
            try:
                return date(int(ano), int(mes), int(dia))
            except Exception:
                return None
        match = re.search(r'(\d{1,3})\s*DIAS', validade, re.IGNORECASE)
        if match and data_emissao_date:
            try:
                return data_emissao_date + timedelta(days=int(match.group(1)))
            except Exception:
                return None
    return None


def clean_info(value):
    """Normaliza valores 'Não informado' para None."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        lower = text.lower()
        if lower in ('nao informado', 'não informado', 'nao informado.'):
            return None
    return value


def apply_pdf_data(proposta, dados):
    """Aplica os dados extraídos do PDF na proposta."""
    proposta.data_emissao = clean_info(dados.get('data_emissao')) or proposta.data_emissao
    proposta.validade = clean_info(dados.get('validade')) or proposta.validade
    proposta.instalacao_status = clean_info(dados.get('instalacao_status')) or proposta.instalacao_status
    proposta.qualificacoes_status = clean_info(dados.get('qualificacoes_status')) or proposta.qualificacoes_status
    proposta.treinamento_status = clean_info(dados.get('treinamento_status')) or proposta.treinamento_status
    proposta.garantia_resumo = clean_info(dados.get('garantia_resumo')) or proposta.garantia_resumo
    proposta.garantia_texto = clean_info(dados.get('garantia_texto')) or proposta.garantia_texto
    proposta.tipo = clean_info(dados.get('tipo')) or clean_info(proposta.tipo) or extract_tipo_from_filename(proposta.nome_arquivo_pdf)
    data_emissao_date = parse_date_br(proposta.data_emissao)
    data_vencimento = compute_data_vencimento(proposta.validade, data_emissao_date)
    if data_vencimento:
        proposta.data_vencimento = data_vencimento


def replace_itens(proposta, itens):
    """Substitui os itens associados a uma proposta."""
    ItemProposta.query.filter_by(proposta_id=proposta.id).delete()
    for item_data in itens or []:
        item = ItemProposta(
            proposta_id=proposta.id,
            numero=item_data.get('numero'),
            descricao=item_data.get('descricao'),
            quantidade=item_data.get('quantidade'),
            valor_unitario=item_data.get('valor_unitario'),
            valor_total=item_data.get('valor_total')
        )
        db.session.add(item)


def _reprocess_all_worker():
    with app.app_context():
        propostas = Proposta.query.filter(Proposta.nome_arquivo_pdf.isnot(None)).all()
        reprocess_status['total'] = len(propostas)
        reprocess_status['done'] = 0
        reprocess_status['errors'] = 0
        batch = 0
        for proposta in propostas:
            if not proposta.nome_arquivo_pdf:
                reprocess_status['done'] += 1
                continue
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], proposta.nome_arquivo_pdf)
            if not os.path.exists(pdf_path):
                reprocess_status['done'] += 1
                continue
            try:
                extractor = PropostaExtractor(pdf_path)
                dados = extractor.extract_all()
                if dados:
                    apply_pdf_data(proposta, dados)
                    replace_itens(proposta, dados.get('itens'))
                    batch += 1
            except Exception as e:
                reprocess_status['errors'] += 1
                print(f"Erro reprocessando {proposta.nome_arquivo_pdf}: {e}")
            reprocess_status['done'] += 1
            if batch >= 25:
                db.session.commit()
                batch = 0
        if batch:
            db.session.commit()
        reprocess_status['running'] = False


def split_proposta_id(id_proposta):
    """Retorna (base_id, versao) a partir do ID da proposta."""
    if not id_proposta:
        return None, None
    match = re.match(r'^([A-Z]{2})\.(\d{3,4})([A-Z]?)/(\d{2,4})$', id_proposta.strip(), re.IGNORECASE)
    if not match:
        return id_proposta, ''
    prefixo, numero, versao, ano = match.groups()
    base_id = f"{prefixo.upper()}.{numero}/{ano}"
    return base_id, versao.upper() if versao else ''


def versao_ordem(versao):
    """Converte letra de versão em ordem numérica."""
    if not versao:
        return 0
    if len(versao) == 1 and 'A' <= versao.upper() <= 'Z':
        return ord(versao.upper()) - ord('A') + 1
    return 0


def ensure_schema():
    """Garante colunas novas no SQLite sem migração."""
    with db.engine.connect() as conn:
        result = conn.execute(text("PRAGMA table_info(propostas)"))
        existing = {row[1] for row in result.fetchall()}
        if 'cod_vendedor' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN cod_vendedor VARCHAR(50)"))
        if 'data_vencimento' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN data_vencimento DATE"))
        if 'instalacao_status' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN instalacao_status VARCHAR(30)"))
        if 'qualificacoes_status' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN qualificacoes_status VARCHAR(30)"))
        if 'treinamento_status' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN treinamento_status VARCHAR(30)"))
        if 'garantia_resumo' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN garantia_resumo TEXT"))
        if 'garantia_texto' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN garantia_texto TEXT"))
        if 'tipo' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN tipo VARCHAR(20)"))
        if 'observacoes' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN observacoes VARCHAR(30)"))
        if 'id_proposta_base' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN id_proposta_base VARCHAR(50)"))
        if 'versao' not in existing:
            conn.execute(text("ALTER TABLE propostas ADD COLUMN versao VARCHAR(5)"))
        # Índices para acelerar a listagem/paginação
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_base ON propostas(id_proposta_base)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_import ON propostas(data_importacao)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_obs ON propostas(observacoes)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_cnpj ON propostas(cnpj)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_razao ON propostas(razao_social)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_cod ON propostas(cod_vendedor)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_propostas_id ON propostas(id_proposta)"))

        # Visitas: ajustes de schema
        try:
            result_visitas = conn.execute(text("PRAGMA table_info(visitas)"))
            existing_visitas = {row[1] for row in result_visitas.fetchall()}
            if 'historico' not in existing_visitas and existing_visitas:
                conn.execute(text("ALTER TABLE visitas ADD COLUMN historico TEXT"))
            if existing_visitas and ('data' not in existing_visitas):
                if 'data_hora' in existing_visitas:
                    data_expr = "date(data_hora)"
                elif 'data_criacao' in existing_visitas:
                    data_expr = "date(data_criacao)"
                else:
                    data_expr = "date('now')"
                historico_expr = "historico" if 'historico' in existing_visitas else "NULL"
                data_criacao_expr = "data_criacao" if 'data_criacao' in existing_visitas else "NULL"

                conn.execute(text("PRAGMA foreign_keys=off"))
                conn.execute(text("""
                    CREATE TABLE visitas_new (
                        id INTEGER PRIMARY KEY,
                        cliente_id INTEGER NOT NULL,
                        data DATE NOT NULL,
                        historico TEXT,
                        data_criacao DATETIME,
                        FOREIGN KEY(cliente_id) REFERENCES clientes(id)
                    )
                """))
                conn.execute(text(f"""
                    INSERT INTO visitas_new (id, cliente_id, data, historico, data_criacao)
                    SELECT id, cliente_id, {data_expr}, {historico_expr}, {data_criacao_expr}
                    FROM visitas
                """))
                conn.execute(text("DROP TABLE visitas"))
                conn.execute(text("ALTER TABLE visitas_new RENAME TO visitas"))
                conn.execute(text("PRAGMA foreign_keys=on"))
        except Exception:
            # Tabela pode não existir ainda
            pass

        # Contatos: adicionar setor se faltar
        try:
            result_contatos = conn.execute(text("PRAGMA table_info(contatos)"))
            existing_contatos = {row[1] for row in result_contatos.fetchall()}
            if 'setor' not in existing_contatos and existing_contatos:
                conn.execute(text("ALTER TABLE contatos ADD COLUMN setor VARCHAR(100)"))
        except Exception:
            # Tabela pode não existir ainda
            pass

        # Clientes: adicionar CPM se faltar
        try:
            result_clientes = conn.execute(text("PRAGMA table_info(clientes)"))
            existing_clientes = {row[1] for row in result_clientes.fetchall()}
            if 'cpm_status' not in existing_clientes and existing_clientes:
                conn.execute(text("ALTER TABLE clientes ADD COLUMN cpm_status VARCHAR(20)"))
            if 'cpm_data' not in existing_clientes and existing_clientes:
                conn.execute(text("ALTER TABLE clientes ADD COLUMN cpm_data DATE"))
            if 'regiao' not in existing_clientes and existing_clientes:
                conn.execute(text("ALTER TABLE clientes ADD COLUMN regiao VARCHAR(50)"))
        except Exception:
            pass

# Inicializar banco de dados
db.init_app(app)
with app.app_context():
    db.create_all()
    ensure_schema()


def allowed_file(filename):
    """Verifica se o arquivo é permitido"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    """Página principal - redireciona para upload"""
    return redirect(url_for('listagem'))


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """Página de upload de PDF"""
    if request.method == 'POST':
        # Verificar se arquivo foi enviado
        files = request.files.getlist('pdf_files')
        if not files or (len(files) == 1 and files[0].filename == ''):
            files = request.files.getlist('pdf_file')

        if not files or all(f.filename == '' for f in files):
            flash('Nenhum arquivo foi selecionado', 'error')
            return redirect(request.url)

        if len(files) > app.config['MAX_FILES_PER_UPLOAD']:
            flash(f'Limite de {app.config["MAX_FILES_PER_UPLOAD"]} PDFs por envio.', 'error')
            return redirect(request.url)

        enfileirados = 0
        for file_obj in files:
            if file_obj.filename == '':
                continue
            if not allowed_file(file_obj.filename):
                flash(f'Apenas arquivos PDF são permitidos: {file_obj.filename}', 'error')
                continue

            filename_original = file_obj.filename
            filename = secure_filename(filename_original)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file_obj.save(filepath)
            upload_queue.put((filepath, filename_original, filename))
            enfileirados += 1

        if enfileirados:
            with upload_lock:
                global upload_total, upload_done
                upload_total += enfileirados

        if enfileirados:
            flash(f'{enfileirados} PDF(s) enviados para processamento em background.', 'success')
        else:
            flash('Nenhum PDF válido foi enviado.', 'warning')

        return redirect(url_for('listagem'))
    
    return render_template('upload.html')


@app.route('/listagem')
def listagem():
    """Página de listagem de propostas"""
    # Filtros
    razao_social = request.args.get('razao_social', '').strip()
    cnpj = request.args.get('cnpj', '').strip()
    id_proposta = request.args.get('id_proposta', '').strip()
    cod_vendedor = request.args.get('cod_vendedor', '').strip()
    sort = request.args.get('sort', 'id_proposta').strip()
    order = request.args.get('order', 'asc').strip().lower()
    page = request.args.get('page', '1').strip()
    per_page = request.args.get('per_page', '50').strip()
    try:
        page = max(int(page), 1)
    except Exception:
        page = 1
    try:
        per_page = int(per_page)
    except Exception:
        per_page = 50
    if per_page not in (10, 50, 150):
        per_page = 50
    if order not in ('asc', 'desc'):
        order = 'asc'
    
    # Query base (apenas a proposta atual por base)
    base_expr = func.coalesce(Proposta.id_proposta_base, Proposta.id_proposta)
    subq = db.session.query(
        base_expr.label('base_id'),
        func.max(Proposta.data_importacao).label('max_import')
    ).group_by(base_expr).subquery()

    current_query = Proposta.query.join(
        subq,
        (base_expr == subq.c.base_id) & (Proposta.data_importacao == subq.c.max_import)
    )

    # Aplicar filtros
    if razao_social:
        current_query = current_query.filter(Proposta.razao_social.ilike(f'%{razao_social}%'))
    if cnpj:
        current_query = current_query.filter(Proposta.cnpj.ilike(f'%{cnpj}%'))
    if id_proposta:
        current_query = current_query.filter(Proposta.id_proposta.ilike(f'%{id_proposta}%'))
    if cod_vendedor:
        current_query = current_query.filter(Proposta.cod_vendedor.ilike(f'%{cod_vendedor}%'))

    # Ordenacao
    if sort == 'data_emissao':
        emissao_expr = func.substr(Proposta.data_emissao, 7, 4)
        emissao_expr = emissao_expr.op('||')(literal('-')).op('||')(func.substr(Proposta.data_emissao, 4, 2)).op('||')(literal('-')).op('||')(func.substr(Proposta.data_emissao, 1, 2))
        order_col = func.coalesce(emissao_expr, '')
    elif sort == 'data_vencimento':
        order_col = func.coalesce(Proposta.data_vencimento, date.min)
    elif sort == 'status':
        order_col = func.lower(func.coalesce(Proposta.observacoes, ''))
    elif sort == 'cnpj':
        order_col = func.lower(func.coalesce(Proposta.cnpj, ''))
    elif sort == 'razao_social':
        order_col = func.lower(func.coalesce(Proposta.razao_social, ''))
    elif sort == 'cod_vendedor':
        order_col = func.lower(func.coalesce(Proposta.cod_vendedor, ''))
    else:
        order_col = func.lower(func.coalesce(Proposta.id_proposta, ''))

    if order == 'desc':
        current_query = current_query.order_by(order_col.desc())
    else:
        current_query = current_query.order_by(order_col.asc())

    count_query = current_query.order_by(None)
    total_propostas = count_query.count()
    total_pages = max(ceil(total_propostas / per_page), 1)
    if page > total_pages:
        page = total_pages
    start_idx = (page - 1) * per_page

    current_propostas = current_query.offset(start_idx).limit(per_page).all()

    base_ids = [p.id_proposta_base or p.id_proposta for p in current_propostas]
    grupos_lista = []
    propostas_para_backfill = []

    if base_ids:
        all_versions = Proposta.query.filter(base_expr.in_(base_ids)).all()
        propostas_para_backfill = all_versions
        grupos = {}
        for proposta in all_versions:
            base_id = proposta.id_proposta_base or proposta.id_proposta
            grupos.setdefault(base_id, []).append(proposta)

        for current in current_propostas:
            base_id = current.id_proposta_base or current.id_proposta
            itens = grupos.get(base_id, [])
            itens_ordenados = sorted(
                itens,
                key=lambda p: (versao_ordem(p.versao or ''), p.data_importacao or datetime.min)
            )
            versions = [p for p in reversed(itens_ordenados) if p.id != current.id]
            grupos_lista.append({'current': current, 'versions': versions})

    hoje = date.today()
    limite_vencendo = hoje + timedelta(days=7)
    total_vencidas = 0
    alterou = False
    for proposta in propostas_para_backfill:
        # Backfill de data de vencimento se faltar
        if not proposta.data_vencimento and proposta.data_emissao:
            data_emissao_date = parse_date_br(proposta.data_emissao)
            data_vencimento = compute_data_vencimento(proposta.validade, data_emissao_date)
            if not data_vencimento and data_emissao_date:
                data_vencimento = data_emissao_date + timedelta(days=30)
            if data_vencimento:
                proposta.data_vencimento = data_vencimento
                alterou = True

        # Backfill de cod pelo nome do arquivo
        if not proposta.cod_vendedor and proposta.nome_arquivo_pdf:
            cod = extract_cod_from_filename(proposta.nome_arquivo_pdf)
            if cod:
                proposta.cod_vendedor = cod
                alterou = True
        if proposta.razao_social:
            razao_upper = proposta.razao_social.upper()
            if proposta.razao_social != razao_upper:
                proposta.razao_social = razao_upper
                alterou = True

        # Backfill leve para tipo (sem abrir PDF)
        if (proposta.tipo is None or clean_info(proposta.tipo) is None) and proposta.nome_arquivo_pdf:
            tipo = extract_tipo_from_filename(proposta.nome_arquivo_pdf)
            if tipo:
                proposta.tipo = tipo
                alterou = True

        # Flags para status de vencimento
        proposta.vencida = False
        proposta.vencendo = False
        if proposta.data_vencimento:
            if proposta.data_vencimento < hoje:
                proposta.vencida = True
                total_vencidas += 1
            elif proposta.data_vencimento <= limite_vencendo:
                proposta.vencendo = True

        # Backfill base/versao com base nos 3 primeiros dígitos do arquivo
        base_id = extract_base_id_from_filename(proposta.nome_arquivo_pdf) if proposta.nome_arquivo_pdf else None
        versao = extract_version_from_filename(proposta.nome_arquivo_pdf) if proposta.nome_arquivo_pdf else None
        if base_id and proposta.id_proposta_base != base_id:
            proposta.id_proposta_base = base_id
            proposta.versao = versao or proposta.versao
            alterou = True
        elif not proposta.id_proposta_base:
            base_id_fallback, versao_fallback = split_proposta_id(proposta.id_proposta)
            proposta.id_proposta_base = base_id_fallback
            proposta.versao = proposta.versao or versao_fallback
            alterou = True

        if proposta.vencida:
            if proposta.observacoes in (None, '', 'Em negociação', 'Vencida'):
                if proposta.observacoes != 'Vencida':
                    proposta.observacoes = 'Vencida'
                    alterou = True
        elif not proposta.observacoes:
            proposta.observacoes = 'Em negociação'
            alterou = True

    if alterou:
        db.session.commit()

    totals = count_query.with_entities(
        func.sum(case((Proposta.observacoes == 'Ganha', 1), else_=0)).label('ganhas'),
        func.sum(case((Proposta.observacoes == 'Perdida', 1), else_=0)).label('perdidas'),
        func.sum(case((Proposta.observacoes == 'Em negociação', 1), else_=0)).label('abertas'),
        func.sum(case((Proposta.observacoes == 'Vencida', 1), else_=0)).label('vencidas')
    ).first()
    total_ganhas = totals.ganhas or 0
    total_perdidas = totals.perdidas or 0
    total_abertas = totals.abertas or 0
    total_vencidas_dashboard = totals.vencidas or 0
    total_vencidas = total_vencidas_dashboard

    return render_template('listagem.html', 
                         grupos=grupos_lista,
                         filtros={
                             'razao_social': razao_social,
                             'cnpj': cnpj,
                             'id_proposta': id_proposta,
                             'cod_vendedor': cod_vendedor,
                             'sort': sort,
                             'order': order,
                             'per_page': per_page
                         },
                         total_vencidas=total_vencidas,
                         total_propostas=total_propostas,
                         total_ganhas=total_ganhas,
                         total_perdidas=total_perdidas,
                         total_abertas=total_abertas,
                         total_vencidas_dashboard=total_vencidas_dashboard,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages,
                         sort=sort,
                         order=order)


@app.route('/api/upload_status')
def upload_status():
    """Status da fila de importação"""
    with upload_lock:
        global upload_total, upload_done
        total = upload_total
        done = upload_done
        pending = max(total - done, 0)
        if total and pending == 0:
            # Reset to avoid oscillating progress bar after completion
            upload_total = 0
            upload_done = 0
    percent = int((done / total) * 100) if total else 0
    return jsonify({
        'total': total,
        'done': done,
        'pending': pending,
        'percent': percent,
        'completed': True if total and pending == 0 else False
    })


@app.route('/clientes')
def clientes():
    """Página de listagem de clientes"""
    clientes = Cliente.query.order_by(Cliente.nome.asc()).all()
    return render_template('clientes_listagem.html', clientes=clientes)


@app.route('/relatorio')
def relatorio():
    """Página de relatórios"""
    clientes = Cliente.query.order_by(Cliente.nome.asc()).all()
    return render_template('relatorio.html', clientes=clientes)


def _autosize_sheet(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


@app.route('/relatorio/export', methods=['POST'])
def relatorio_export():
    """Exporta relatório em XLSX"""
    cliente_id = request.form.get('cliente_id', '').strip()
    if not cliente_id or not cliente_id.isdigit():
        flash('Selecione um cliente para exportar.', 'error')
        return redirect(url_for('relatorio'))

    cliente = Cliente.query.get_or_404(int(cliente_id))
    contatos = Contato.query.filter_by(cliente_id=cliente.id).order_by(Contato.nome.asc()).all()
    visitas = Visita.query.filter_by(cliente_id=cliente.id).order_by(Visita.data.desc()).all()
    equipamentos = Equipamento.query.filter_by(cliente_id=cliente.id).order_by(Equipamento.nome.asc()).all()
    cnpj_normalizado = cliente.cnpj_normalizado or normalize_cnpj(cliente.cnpj)
    propostas = Proposta.query.filter(
        (Proposta.cnpj == cliente.cnpj) |
        (func.replace(func.replace(func.replace(Proposta.cnpj, '.', ''), '/', ''), '-', '') == cnpj_normalizado)
    ).order_by(Proposta.id_proposta.asc()).all()

    wb = Workbook()

    ws_info = wb.active
    ws_info.title = 'Cliente'
    ws_info.append(['Campo', 'Valor'])
    ws_info.append(['Nome', cliente.nome])
    ws_info.append(['CNPJ', cliente.cnpj])
    ws_info.append(['Contato principal', cliente.contato or ''])
    ws_info.append(['Email principal', cliente.email or ''])
    ws_info.append(['Telefone principal', cliente.telefone or ''])
    ws_info.append(['Endereço', cliente.endereco or ''])
    ws_info.append(['Região', cliente.regiao or ''])
    ws_info.append(['CPM Status', cliente.cpm_status or ''])
    if cliente.cpm_data:
        ws_info.append(['CPM Data', cliente.cpm_data.strftime('%d/%m/%Y')])
    _autosize_sheet(ws_info)

    ws_contatos = wb.create_sheet('Contatos')
    ws_contatos.append(['Nome', 'Email', 'Telefone', 'Cargo', 'Setor'])
    for c in contatos:
        ws_contatos.append([c.nome, c.email or '', c.telefone or '', c.cargo or '', c.setor or ''])
    _autosize_sheet(ws_contatos)

    ws_equip = wb.create_sheet('Parque instalado')
    ws_equip.append(['Equipamento', 'Quantidade', 'Marca', 'Modelo', 'Ano'])
    for e in equipamentos:
        ws_equip.append([
            e.nome,
            e.quantidade if e.quantidade is not None else '',
            e.marca or '',
            e.modelo or '',
            e.ano_instalacao if e.ano_instalacao is not None else ''
        ])
    _autosize_sheet(ws_equip)

    ws_visitas = wb.create_sheet('Visitas')
    ws_visitas.append(['Data', 'Histórico'])
    for v in visitas:
        ws_visitas.append([
            v.data.strftime('%d/%m/%Y') if v.data else '',
            v.historico or ''
        ])
    _autosize_sheet(ws_visitas)

    ws_propostas = wb.create_sheet('Propostas')
    ws_propostas.append(['ID Proposta', 'Razão Social', 'Data Emissão', 'Valor Total', 'Cod Vendedor'])
    for p in propostas:
        ws_propostas.append([
            p.id_proposta or '',
            p.razao_social or '',
            p.data_emissao or '',
            p.valor_total or '',
            p.cod_vendedor or ''
        ])
    _autosize_sheet(ws_propostas)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_cliente_{cliente.id}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/relatorio/export_db')
def relatorio_export_db():
    """Exporta o banco SQLite completo"""
    db_path = os.path.join(app.instance_path, 'database.db')
    if not os.path.exists(db_path):
        flash('Banco de dados nao encontrado.', 'error')
        return redirect(url_for('relatorio'))
    return send_file(
        db_path,
        mimetype='application/x-sqlite3',
        as_attachment=True,
        download_name='database.db'
    )


@app.route('/clientes/novo', methods=['GET', 'POST'])
def clientes_novo():
    """Cadastro de novos clientes"""
    setores = Setor.query.order_by(Setor.nome.asc()).all()
    regioes = Regiao.query.order_by(Regiao.nome.asc()).all()
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        contato = request.form.get('contato', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        endereco = request.form.get('endereco', '').strip()
        cpm_status = cliente.cpm_status
        cpm_data = cliente.cpm_data
        if 'cpm_status' in request.form or 'cpm_data' in request.form:
            cpm_status = request.form.get('cpm_status', '').strip()
            cpm_data_str = request.form.get('cpm_data', '').strip()
            cpm_data = parse_date_iso(cpm_data_str)
        regiao = request.form.get('regiao', '').strip()

        if not nome or not cnpj:
            flash('Nome e CNPJ são obrigatórios.', 'error')
            return redirect(request.url)

        cnpj_normalizado = normalize_cnpj(cnpj)
        existente = Cliente.query.filter(
            (Cliente.cnpj == cnpj) | (Cliente.cnpj_normalizado == cnpj_normalizado)
        ).first()
        if existente:
            flash('Já existe um cliente cadastrado com esse CNPJ.', 'warning')
            return redirect(url_for('cliente_detalhes', id=existente.id))

        cliente = Cliente(
            nome=nome,
            cnpj=cnpj,
            cnpj_normalizado=cnpj_normalizado,
            contato=contato or None,
            email=email or None,
            telefone=telefone or None,
            endereco=endereco or None,
            setor=None,
            cpm_status=cpm_status or None,
            cpm_data=cpm_data,
            regiao=regiao or None
        )
        db.session.add(cliente)
        db.session.flush()

        nomes_contatos = request.form.getlist('contato_nome[]')
        emails_contatos = request.form.getlist('contato_email[]')
        telefones_contatos = request.form.getlist('contato_telefone[]')
        cargos_contatos = request.form.getlist('contato_cargo[]')
        setores_contatos = request.form.getlist('contato_setor[]')

        for idx, nome_contato in enumerate(nomes_contatos):
            nome_contato = nome_contato.strip()
            if not nome_contato:
                continue
            contato_item = Contato(
                cliente_id=cliente.id,
                nome=nome_contato,
                email=(emails_contatos[idx].strip() if idx < len(emails_contatos) else '') or None,
                telefone=(telefones_contatos[idx].strip() if idx < len(telefones_contatos) else '') or None,
                cargo=(cargos_contatos[idx].strip() if idx < len(cargos_contatos) else '') or None,
                setor=(setores_contatos[idx].strip() if idx < len(setores_contatos) else '') or None
            )
            db.session.add(contato_item)
        db.session.commit()
        flash('Cliente criado com sucesso!', 'success')
        return redirect(url_for('cliente_detalhes', id=cliente.id))

    return render_template('cliente_novo.html', setores=setores, regioes=regioes)


@app.route('/setores/novo', methods=['POST'])
def setores_novo():
    """Cadastro rápido de setor"""
    nome = request.form.get('setor_nome', '').strip()
    if not nome:
        flash('Informe o nome do setor.', 'error')
        return redirect(request.referrer or url_for('clientes_novo'))

    existente = Setor.query.filter_by(nome=nome).first()
    if existente:
        flash('Este setor já está cadastrado.', 'warning')
        return redirect(request.referrer or url_for('clientes_novo'))

    setor = Setor(nome=nome)
    db.session.add(setor)
    db.session.commit()
    flash('Setor adicionado com sucesso!', 'success')
    return redirect(request.referrer or url_for('clientes_novo'))


@app.route('/regioes/novo', methods=['POST'])
def regioes_novo():
    """Cadastro rápido de região"""
    nome = request.form.get('regiao_nome', '').strip()
    if not nome:
        flash('Informe o nome da região.', 'error')
        return redirect(request.referrer or url_for('clientes_novo'))

    existente = Regiao.query.filter_by(nome=nome).first()
    if existente:
        flash('Esta região já está cadastrada.', 'warning')
        return redirect(request.referrer or url_for('clientes_novo'))

    regiao = Regiao(nome=nome)
    db.session.add(regiao)
    db.session.commit()
    flash('Região adicionada com sucesso!', 'success')
    return redirect(request.referrer or url_for('clientes_novo'))


@app.route('/clientes/<int:id>')
def cliente_detalhes(id):
    """Detalhes do cliente e propostas vinculadas"""
    cliente = Cliente.query.get_or_404(id)
    cnpj_normalizado = cliente.cnpj_normalizado or normalize_cnpj(cliente.cnpj)
    propostas = Proposta.query.filter(
        (Proposta.cnpj == cliente.cnpj) |
        (func.replace(func.replace(func.replace(Proposta.cnpj, '.', ''), '/', ''), '-', '') == cnpj_normalizado)
    ).order_by(Proposta.id_proposta.asc()).all()
    visitas = Visita.query.filter_by(cliente_id=cliente.id).order_by(Visita.data.desc()).all()
    setores = Setor.query.order_by(Setor.nome.asc()).all()
    contatos = Contato.query.filter_by(cliente_id=cliente.id).order_by(Contato.nome.asc()).all()
    equipamentos = Equipamento.query.filter_by(cliente_id=cliente.id).order_by(Equipamento.nome.asc()).all()

    return render_template('cliente_detalhes.html', cliente=cliente, propostas=propostas, visitas=visitas,
                           contatos=contatos, setores=setores, equipamentos=equipamentos)


@app.route('/clientes/<int:id>/cpm', methods=['POST'])
def cliente_cpm_atualizar(id):
    """Atualiza CPM do cliente na tela de detalhes"""
    cliente = Cliente.query.get_or_404(id)
    cpm_status = request.form.get('cpm_status', '').strip()
    cpm_data_str = request.form.get('cpm_data', '').strip()
    cpm_data = parse_date_iso(cpm_data_str)

    cliente.cpm_status = cpm_status or None
    cliente.cpm_data = cpm_data

    db.session.commit()
    flash('CPM atualizado com sucesso!', 'success')
    return redirect(url_for('cliente_detalhes', id=cliente.id))


@app.route('/clientes/<int:id>/editar', methods=['GET', 'POST'])
def cliente_editar(id):
    """Editar cliente"""
    cliente = Cliente.query.get_or_404(id)

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        contato = request.form.get('contato', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        endereco = request.form.get('endereco', '').strip()
        cpm_status = request.form.get('cpm_status', '').strip()
        cpm_data_str = request.form.get('cpm_data', '').strip()
        cpm_data = parse_date_iso(cpm_data_str)
        regiao = request.form.get('regiao', '').strip()

        if not nome or not cnpj:
            flash('Nome e CNPJ são obrigatórios.', 'error')
            return redirect(request.url)

        cnpj_normalizado = normalize_cnpj(cnpj)
        existente = Cliente.query.filter(
            ((Cliente.cnpj == cnpj) | (Cliente.cnpj_normalizado == cnpj_normalizado)) &
            (Cliente.id != cliente.id)
        ).first()
        if existente:
            flash('Já existe outro cliente com esse CNPJ.', 'warning')
            return redirect(url_for('cliente_editar', id=cliente.id))

        cliente.nome = nome
        cliente.cnpj = cnpj
        cliente.cnpj_normalizado = cnpj_normalizado
        cliente.contato = contato or None
        cliente.email = email or None
        cliente.telefone = telefone or None
        cliente.endereco = endereco or None
        cliente.setor = None
        cliente.cpm_status = cpm_status or None
        cliente.cpm_data = cpm_data
        cliente.regiao = regiao or None

        db.session.commit()
        flash('Cliente atualizado com sucesso!', 'success')
        return redirect(url_for('cliente_detalhes', id=cliente.id))

    regioes = Regiao.query.order_by(Regiao.nome.asc()).all()
    return render_template('cliente_editar.html', cliente=cliente, regioes=regioes)


@app.route('/clientes/<int:id>/contatos/novo', methods=['POST'])
def contatos_novo(id):
    """Cria um novo contato para o cliente"""
    cliente = Cliente.query.get_or_404(id)
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip()
    telefone = request.form.get('telefone', '').strip()
    cargo = request.form.get('cargo', '').strip()
    setor = request.form.get('setor', '').strip()

    if not nome:
        flash('Informe o nome do contato.', 'error')
        return redirect(url_for('cliente_detalhes', id=cliente.id))

    contato = Contato(
        cliente_id=cliente.id,
        nome=nome,
        email=email or None,
        telefone=telefone or None,
        cargo=cargo or None,
        setor=setor or None
    )
    db.session.add(contato)
    db.session.commit()
    flash('Contato adicionado com sucesso!', 'success')
    return redirect(url_for('cliente_detalhes', id=cliente.id))


@app.route('/contatos/<int:id>/editar', methods=['GET', 'POST'])
def contato_editar(id):
    """Editar contato"""
    contato = Contato.query.get_or_404(id)
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        cargo = request.form.get('cargo', '').strip()
        setor = request.form.get('setor', '').strip()

        if not nome:
            flash('Informe o nome do contato.', 'error')
            return redirect(request.url)

        contato.nome = nome
        contato.email = email or None
        contato.telefone = telefone or None
        contato.cargo = cargo or None
        contato.setor = setor or None
        db.session.commit()
        flash('Contato atualizado com sucesso!', 'success')
        return redirect(url_for('cliente_detalhes', id=contato.cliente_id))

    return render_template('contato_editar.html', contato=contato)


@app.route('/contatos/<int:id>/deletar', methods=['POST'])
def contato_deletar(id):
    """Deleta contato"""
    contato = Contato.query.get_or_404(id)
    cliente_id = contato.cliente_id
    try:
        db.session.delete(contato)
        db.session.commit()
        flash('Contato deletado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar contato: {str(e)}', 'error')
    return redirect(url_for('cliente_detalhes', id=cliente_id))


@app.route('/clientes/<int:id>/equipamentos/novo', methods=['POST'])
def equipamentos_novo(id):
    """Cria equipamentos para o cliente"""
    cliente = Cliente.query.get_or_404(id)
    nomes = request.form.getlist('equipamento_nome[]')
    quantidades = request.form.getlist('equipamento_quantidade[]')
    marcas = request.form.getlist('equipamento_marca[]')
    modelos = request.form.getlist('equipamento_modelo[]')
    anos = request.form.getlist('equipamento_ano[]')

    adicionou = False
    for idx, nome in enumerate(nomes):
        nome = nome.strip()
        if not nome:
            continue
        quantidade = (quantidades[idx].strip() if idx < len(quantidades) else '')
        ano = (anos[idx].strip() if idx < len(anos) else '')
        equipamento = Equipamento(
            cliente_id=cliente.id,
            nome=nome,
            quantidade=int(quantidade) if quantidade.isdigit() else None,
            marca=(marcas[idx].strip() if idx < len(marcas) else '') or None,
            modelo=(modelos[idx].strip() if idx < len(modelos) else '') or None,
            ano_instalacao=int(ano) if ano.isdigit() else None
        )
        db.session.add(equipamento)
        adicionou = True

    if not adicionou:
        flash('Informe ao menos o nome do equipamento.', 'error')
        return redirect(url_for('cliente_detalhes', id=cliente.id))

    db.session.commit()
    flash('Equipamento(s) adicionados com sucesso!', 'success')
    return redirect(url_for('cliente_detalhes', id=cliente.id))


@app.route('/equipamentos/<int:id>/editar', methods=['GET', 'POST'])
def equipamento_editar(id):
    """Editar equipamento"""
    equipamento = Equipamento.query.get_or_404(id)
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        quantidade = request.form.get('quantidade', '').strip()
        marca = request.form.get('marca', '').strip()
        modelo = request.form.get('modelo', '').strip()
        ano = request.form.get('ano', '').strip()

        if not nome:
            flash('Informe o nome do equipamento.', 'error')
            return redirect(request.url)

        equipamento.nome = nome
        equipamento.quantidade = int(quantidade) if quantidade.isdigit() else None
        equipamento.marca = marca or None
        equipamento.modelo = modelo or None
        equipamento.ano_instalacao = int(ano) if ano.isdigit() else None
        db.session.commit()
        flash('Equipamento atualizado com sucesso!', 'success')
        return redirect(url_for('cliente_detalhes', id=equipamento.cliente_id))

    return render_template('equipamento_editar.html', equipamento=equipamento)


@app.route('/equipamentos/<int:id>/deletar', methods=['POST'])
def equipamento_deletar(id):
    """Deleta equipamento"""
    equipamento = Equipamento.query.get_or_404(id)
    cliente_id = equipamento.cliente_id
    try:
        db.session.delete(equipamento)
        db.session.commit()
        flash('Equipamento deletado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar equipamento: {str(e)}', 'error')
    return redirect(url_for('cliente_detalhes', id=cliente_id))


@app.route('/clientes/<int:id>/deletar', methods=['POST'])
def cliente_deletar(id):
    """Deleta um cliente e seus registros relacionados"""
    cliente = Cliente.query.get_or_404(id)
    try:
        Contato.query.filter_by(cliente_id=cliente.id).delete()
        Equipamento.query.filter_by(cliente_id=cliente.id).delete()
        Visita.query.filter_by(cliente_id=cliente.id).delete()
        db.session.delete(cliente)
        db.session.commit()
        flash('Cliente deletado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar cliente: {str(e)}', 'error')
    return redirect(url_for('clientes'))


@app.route('/clientes/<int:id>/visitas/nova', methods=['POST'])
def visitas_nova(id):
    """Cria uma nova visita para o cliente"""
    cliente = Cliente.query.get_or_404(id)
    data_str = request.form.get('data', '').strip()
    historico = request.form.get('historico', '').strip()

    if not data_str:
        flash('Informe a data da visita.', 'error')
        return redirect(url_for('cliente_detalhes', id=cliente.id))

    try:
        data_visita = datetime.strptime(data_str, '%Y-%m-%d').date()
    except Exception:
        flash('Formato de data inválido.', 'error')
        return redirect(url_for('cliente_detalhes', id=cliente.id))

    visita = Visita(cliente_id=cliente.id, data=data_visita, historico=historico or None)
    db.session.add(visita)
    db.session.commit()
    flash('Visita registrada com sucesso!', 'success')
    return redirect(url_for('cliente_detalhes', id=cliente.id))


@app.route('/detalhes/<int:id>')
def detalhes(id):
    """Página de detalhes de uma proposta"""
    proposta = Proposta.query.get_or_404(id)
    itens = ItemProposta.query.filter_by(proposta_id=id).all()
    total_itens = 0.0
    for item in itens:
        if not item.valor_total:
            continue
        try:
            valor = item.valor_total.replace('.', '').replace(',', '.')
            total_itens += float(valor)
        except Exception:
            continue
    return render_template('detalhes.html', proposta=proposta, itens=itens, total_itens=total_itens)


@app.route('/uploads/<path:filename>')
def visualizar_pdf(filename):
    """Serve PDFs enviados para visualização no navegador."""
    safe_name = secure_filename(filename)
    if not safe_name:
        return redirect(url_for('listagem'))
    return send_from_directory(app.config['UPLOAD_FOLDER'], safe_name, as_attachment=False)

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    """Editar proposta"""
    proposta = Proposta.query.get_or_404(id)
    if request.method == 'POST':
        proposta.id_proposta = request.form.get('id_proposta', '').strip() or proposta.id_proposta
        proposta.data_emissao = request.form.get('data_emissao', '').strip() or proposta.data_emissao
        data_vencimento_str = request.form.get('data_vencimento', '').strip()
        proposta.data_vencimento = parse_date_br(data_vencimento_str) if data_vencimento_str else proposta.data_vencimento
        proposta.validade = request.form.get('validade', '').strip() or proposta.validade
        proposta.cod_vendedor = request.form.get('cod_vendedor', '').strip() or proposta.cod_vendedor

        razao_social = request.form.get('razao_social', '').strip()
        proposta.razao_social = razao_social.upper() if razao_social else proposta.razao_social
        proposta.nome_fantasia = request.form.get('nome_fantasia', '').strip() or proposta.nome_fantasia
        proposta.cnpj = request.form.get('cnpj', '').strip() or proposta.cnpj
        proposta.telefone = request.form.get('telefone', '').strip() or proposta.telefone
        proposta.celular = request.form.get('celular', '').strip() or proposta.celular
        proposta.email = request.form.get('email', '').strip() or proposta.email
        proposta.pessoa_contato = request.form.get('pessoa_contato', '').strip() or proposta.pessoa_contato

        proposta.instalacao_status = request.form.get('instalacao_status', '').strip() or proposta.instalacao_status
        proposta.qualificacoes_status = request.form.get('qualificacoes_status', '').strip() or proposta.qualificacoes_status
        proposta.treinamento_status = request.form.get('treinamento_status', '').strip() or proposta.treinamento_status
        proposta.garantia_resumo = request.form.get('garantia_resumo', '').strip() or proposta.garantia_resumo
        observacoes = request.form.get('observacoes', '').strip()
        proposta.observacoes = observacoes if observacoes else proposta.observacoes

        db.session.commit()
        flash('Proposta atualizada com sucesso!', 'success')
        return redirect(url_for('detalhes', id=proposta.id))

    return render_template('editar.html', proposta=proposta)


@app.route('/atualizar_cod/<int:id>', methods=['POST'])
def atualizar_cod(id):
    """Atualiza o código do vendedor manualmente"""
    proposta = Proposta.query.get_or_404(id)
    cod = request.form.get('cod_vendedor', '').strip()
    proposta.cod_vendedor = cod if cod else None
    db.session.commit()
    flash('Código do vendedor atualizado com sucesso!', 'success')
    return redirect(url_for('listagem'))


@app.route('/atualizar_observacoes/<int:id>', methods=['POST'])
def atualizar_observacoes(id):
    """Atualiza observações da proposta na listagem"""
    proposta = Proposta.query.get_or_404(id)
    observacoes = request.form.get('observacoes', '').strip()
    proposta.observacoes = observacoes if observacoes else None
    db.session.commit()
    flash('Observações atualizadas com sucesso!', 'success')
    return redirect(url_for('listagem'))


@app.route('/reprocessar_pdf/<int:id>', methods=['POST'])
def reprocessar_pdf(id):
    """Reprocessa o PDF da proposta e atualiza campos extraídos."""
    proposta = Proposta.query.get_or_404(id)
    if not proposta.nome_arquivo_pdf:
        flash('Proposta sem PDF associado.', 'warning')
        return redirect(url_for('listagem'))

    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], proposta.nome_arquivo_pdf)
    if not os.path.exists(pdf_path):
        flash('Arquivo PDF não encontrado no servidor.', 'danger')
        return redirect(url_for('listagem'))

    try:
        extractor = PropostaExtractor(pdf_path)
        dados = extractor.extract_all()
        if not dados:
            flash('Não foi possível extrair informações do PDF.', 'warning')
            return redirect(url_for('listagem'))

        apply_pdf_data(proposta, dados)
        replace_itens(proposta, dados.get('itens'))

        db.session.commit()
        flash('PDF reprocessado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reprocessar PDF: {str(e)}', 'danger')

    return redirect(url_for('listagem'))


@app.route('/reprocessar_todos', methods=['POST'])
def reprocessar_todos():
    """Reprocessa todos os PDFs disponíveis e atualiza campos extraídos."""
    if reprocess_status.get('running'):
        flash('Reprocessamento já em andamento.', 'warning')
        return redirect(url_for('listagem'))

    reprocess_status['running'] = True
    thread = threading.Thread(target=_reprocess_all_worker, daemon=True)
    thread.start()
    flash('Reprocessamento iniciado em segundo plano. Aguarde alguns minutos.', 'success')
    return redirect(url_for('listagem'))


@app.route('/deletar/<int:id>', methods=['POST'])
def deletar(id):
    """Deletar uma proposta"""
    proposta = Proposta.query.get_or_404(id)
    
    try:
        # Deletar arquivo PDF
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], proposta.nome_arquivo_pdf)
        if os.path.exists(filepath):
            os.remove(filepath)
        
        # Deletar do banco (itens serão deletados em cascata)
        db.session.delete(proposta)
        db.session.commit()
        
        flash(f'Proposta {proposta.id_proposta} deletada com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao deletar proposta: {str(e)}', 'error')
        db.session.rollback()
    
    return redirect(url_for('listagem'))


@app.route('/deletar_multiplos', methods=['POST'])
def deletar_multiplos():
    """Deletar múltiplas propostas selecionadas."""
    ids = request.form.getlist('ids')
    if not ids:
        flash('Selecione pelo menos uma proposta para excluir.', 'warning')
        return redirect(url_for('listagem'))

    try:
        ids_int = [int(i) for i in ids]
    except Exception:
        flash('IDs inválidos para exclusão.', 'error')
        return redirect(url_for('listagem'))

    try:
        propostas = Proposta.query.filter(Proposta.id.in_(ids_int)).all()
        for proposta in propostas:
            if proposta.nome_arquivo_pdf:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], proposta.nome_arquivo_pdf)
                if os.path.exists(filepath):
                    os.remove(filepath)
            db.session.delete(proposta)
        db.session.commit()
        flash(f'{len(propostas)} proposta(s) excluída(s) com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir propostas: {str(e)}', 'error')

    return redirect(url_for('listagem'))


@app.route('/api/propostas')
def api_propostas():
    """API para listar propostas (JSON)"""
    propostas = Proposta.query.all()
    return jsonify([p.to_dict() for p in propostas])


@app.route('/api/proposta/<int:id>')
def api_proposta(id):
    """API para obter detalhes de uma proposta (JSON)"""
    proposta = Proposta.query.get_or_404(id)
    itens = ItemProposta.query.filter_by(proposta_id=id).all()
    
    resultado = proposta.to_dict()
    resultado['itens'] = [item.to_dict() for item in itens]
    
    return jsonify(resultado)


if __name__ == '__main__':
    # Criar diretório de uploads se não existir
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    print("Banco de dados inicializado!")
    # Rodar aplicação
    app.run(debug=True, host='0.0.0.0', port=2020)
